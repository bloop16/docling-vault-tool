"""Docling-basierte Batch-Konvertierung (PDF/DOCX/XLSX/PPTX -> Markdown).

Kernlogik fuer die Umwandlung eines grossen Dokumentenordners in einen
Obsidian-Vault aus strukturierten Markdown-Dateien:

* ``build_converter``      -- baut einen konfigurierten Docling-Converter
* ``discover_files``       -- findet rekursiv alle unterstuetzten Dateien
* ``convert_single_file``  -- konvertiert eine Datei, schreibt .md + Assets
* ``init_worker`` /
  ``convert_file_task``    -- Prozess-Pool-Helfer (ein Converter pro Prozess)

Docling erhaelt die Dokumentstruktur (Ueberschriften, Tabellen) statt reinem
Textbrei und extrahiert eingebettete Bilder als eigene Dateien. Jede erzeugte
``.md``-Datei bekommt YAML-Frontmatter mit ``source``, ``original_path`` und
``assets_folder`` -- Obsidian liest das nativ als Properties.

Das Modul ist bewusst frei von Streamlit-Abhaengigkeiten, damit es sowohl vom
Dashboard (``app_streamlit.py``) als auch direkt per CLI genutzt werden kann::

    python docling_worker.py --input /pfad/zu/quellen --output /pfad/zum/vault
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import time
import traceback
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

# Von Docling nativ unterstuetzte Eingabeformate, die dieses Tool anfasst.
SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".xlsx",
    ".pptx",
    ".html",
    ".htm",
    ".md",
    # Bilder: laufen durch die PDF-Pipeline (OCR-Einstellungen greifen).
    ".png",
    ".jpg",
    ".jpeg",
    ".tif",
    ".tiff",
    ".webp",
    # Weitere Docling-Formate (Basisunterstuetzung).
    ".csv",
    ".adoc",
    ".asciidoc",
    ".eml",
    ".epub",
}

# Teilmenge: Bildformate (fuer die OCR-Warnung im Integrationsplan).
IMAGE_INPUT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp"}


def hash_file(path: os.PathLike | str, chunk: int = 1 << 20) -> str:
    """SHA-256 ueber den Dateiinhalt (gestreamt)."""
    import hashlib

    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for block in iter(lambda: fh.read(chunk), b""):
            h.update(block)
    return h.hexdigest()


def find_duplicate_files(
    files: Iterable[os.PathLike | str],
) -> dict[str, list[Path]]:
    """Gruppen inhaltsgleicher Dateien: ``{sha256: [pfade...]}``.

    Nur Gruppen mit >= 2 Dateien werden geliefert. Billig gehalten:
    zunaechst nach Dateigroesse vorgruppiert, gehasht werden nur Dateien,
    deren Groesse mehrfach vorkommt.
    """
    by_size: dict[int, list[Path]] = {}
    for f in files:
        p = Path(f)
        try:
            by_size.setdefault(p.stat().st_size, []).append(p)
        except OSError:
            continue
    groups: dict[str, list[Path]] = {}
    for candidates in by_size.values():
        if len(candidates) < 2:
            continue
        for p in candidates:
            try:
                groups.setdefault(hash_file(p), []).append(p)
            except OSError:
                continue
    return {h: paths for h, paths in groups.items() if len(paths) >= 2}


@dataclass
class ConverterConfig:
    """Konfiguration fuer den Docling-Converter.

    OCR ist standardmaessig aus (``do_ocr=False``) -- nur fuer gescannte PDFs
    ohne Textlayer gezielt aktivieren, da deutlich langsamer.

    ``on_success`` steuert, was nach erfolgreicher Konvertierung mit der
    Originaldatei passiert: ``"keep"`` (Default, nichts), ``"archive"``
    (nach ``archive_dir`` verschieben, Struktur bleibt erhalten) oder
    ``"delete"`` (Original loeschen -- unwiderruflich).
    """

    do_ocr: bool = False
    # OCR-Engine: "easyocr" (Standard; Modelle von GitHub/JaidedAI),
    # "tesseract" (lokal installiertes Tesseract, Sprachcodes wie "deu,eng")
    # oder "rapidocr" (Docling-Default; laedt PP-OCR-Modelle von
    # modelscope.cn -- in vielen Netzen blockiert).
    ocr_engine: str = "easyocr"
    ocr_languages: str = "de,en"
    generate_picture_images: bool = True
    images_scale: float = 2.0
    do_table_structure: bool = True
    on_success: str = "keep"
    archive_dir: str | None = None

    # --- Ablage/Integration im Zielordner ---------------------------------
    # notes_subdir:      Unterordner im Ziel, unter dem die .md abgelegt werden
    #                    (""=Wurzel). Erlaubt Einordnung in einen bestehenden Vault.
    # mirror_structure:  Quell-Verzeichnisstruktur im Ziel spiegeln.
    # attachments_mode:  "central" -> ein Anhang-Ordner (attachments_subdir),
    #                    "adjacent" -> Anhang-Ordner direkt neben der Notiz.
    # attachments_subdir Name des zentralen Anhang-Ordners (Vault-Konvention).
    # add_frontmatter:   YAML-Properties voranstellen (falls der Vault sie nutzt).
    notes_subdir: str = ""
    mirror_structure: bool = True
    attachments_mode: str = "central"
    attachments_subdir: str = "assets"
    add_frontmatter: bool = True

    # --- XLSX-Sonderfall ---------------------------------------------------
    # Arbeitsmappen mit sehr vielen Blaettern koennen Laufzeit und Notizgroesse
    # sprengen. xlsx_sheet_limit begrenzt die Anzahl (0 = alle Blaetter);
    # xlsx_on_limit bestimmt das Verhalten bei Ueberschreitung:
    #   "limit" -> nur die ersten N Blaetter konvertieren (vermerkt im
    #              Frontmatter als sheets_total/sheets_converted)
    #   "skip"  -> Datei ueberspringen (landet im Fehlerprotokoll)
    xlsx_sheet_limit: int = 0
    xlsx_on_limit: str = "limit"


@dataclass
class ConversionResult:
    """Ergebnis einer einzelnen Konvertierung (picklebar fuer den Pool)."""

    source_path: str
    success: bool
    output_path: str | None = None
    assets_folder: str | None = None
    num_images: int = 0
    duration_s: float = 0.0
    # Fehlerinformationen (nur bei success=False gesetzt):
    error: str | None = None           # knappe Zeile: "Typ: Nachricht"
    error_category: str | None = None  # klassifiziert, z. B. "passwortgeschützt"
    error_hint: str | None = None      # Klartext-Handlungshinweis
    error_detail: str | None = None    # voller Traceback (echte Ursache)
    # Nachbearbeitung des Originals (archive/delete):
    moved_to: str | None = None
    # True, wenn mit speicherschonenden Einstellungen konvertiert wurde
    # (Riesenseiten-Erkennung oder automatischer Wiederholungsversuch).
    reduced_mode: bool = False
    # "pypdfium", wenn erst der alternative PDF-Parser die Datei laden konnte.
    pdf_backend: str | None = None
    # Konvertierung ok, aber Archivieren/Loeschen des Originals schlug fehl.
    post_action_error: str | None = None


def _mute_streamlit_bare_mode_warning() -> None:
    """Windows-Spawn-Worker importieren beim Start Streamlit-Teile des
    Dashboard-Elternprozesses; ohne aktiven Script-Lauf loggt Streamlit dann
    "missing ScriptRunContext ... can be ignored when running in bare mode".
    Die Meldung ist laut Streamlit selbst bedeutungslos, wuerde aber pro
    Worker-Prozess im Konsolen-Log auftauchen -- daher gezielt stummschalten
    (die betroffenen Logger geben ausschliesslich diese Warnung aus)."""
    import logging

    for name in (
        "streamlit.runtime.scriptrunner_utils.script_run_context",
        "streamlit.runtime.scriptrunner.script_run_context",
    ):
        logging.getLogger(name).setLevel(logging.ERROR)


def _mute_torch_pin_memory_warning() -> None:
    """torch meldet auf CPU-only-Systemen bei jedem DataLoader "'pin_memory'
    argument is set as true but no accelerator is found". CPU-Betrieb ist
    fuer doc2vault der Normalfall, die Meldung hat keine Konsequenz."""
    import warnings

    warnings.filterwarnings(
        "ignore", message=r".*pin_memory.*no accelerator.*"
    )
    # EasyOCR nutzt quantisierte torch-Module; die Deprecation-Warnung dazu
    # erscheint sonst einmal pro Worker-Prozess.
    warnings.filterwarnings(
        "ignore", message=r".*quantize_per_tensor.*deprecated.*"
    )


def _mute_worker_progress_bars() -> None:
    """Fortschrittsbalken ("Loading weights: 100%|...") sind in parallelen
    Worker-Prozessen nur Log-Rauschen -- pro Worker eine volle tqdm-Zeile."""
    os.environ.setdefault("HF_HUB_DISABLE_PROGRESS_BARS", "1")
    os.environ.setdefault("TQDM_DISABLE", "1")
    os.environ.setdefault("TRANSFORMERS_VERBOSITY", "error")


_mute_streamlit_bare_mode_warning()
_mute_torch_pin_memory_warning()


# Seitenflaeche in PDF-Punkten, ab der eine Seite als "riesig" gilt
# (~A1 und groesser; CAD-Zeichnungen). Solche Seiten sprengen beim Rendern
# mit voller Bildskalierung den Speicher (std::bad_alloc im Preprocess).
HUGE_PAGE_AREA_PT2 = 2_500_000


def _reduced_config(config: ConverterConfig) -> ConverterConfig:
    """Speicherschonende Variante einer Konfiguration.

    Bildskalierung 1.0 und keine Bildextraktion -- die Hauptspeicherfresser
    beim Rendern riesiger Seiten. Alle uebrigen Einstellungen (OCR, Tabellen,
    Ablage) bleiben erhalten.
    """
    from dataclasses import replace

    return replace(config, images_scale=1.0, generate_picture_images=False)


def _is_reduced(config: ConverterConfig) -> bool:
    return config.images_scale <= 1.0 and not config.generate_picture_images


def has_huge_pages(
    path: os.PathLike | str,
    threshold: float = HUGE_PAGE_AREA_PT2,
    max_pages_checked: int = 50,
) -> bool:
    """True, wenn eine PDF Seiten mit riesiger Flaeche enthaelt (billig via
    pypdfium2, nur Seitengroessen, kein Rendern)."""
    try:
        import pypdfium2 as pdfium

        pdf = pdfium.PdfDocument(str(path))
        try:
            for i in range(min(len(pdf), max_pages_checked)):
                page = pdf[i]
                width, height = page.get_size()
                page.close()
                if width * height > threshold:
                    return True
        finally:
            pdf.close()
    except Exception:  # noqa: BLE001 -- im Zweifel normaler Pfad
        return False
    return False


def _make_ocr_options(engine: str, languages: str):
    """Baut die Docling-OCR-Optionen fuer die gewaehlte Engine.

    Standard ist EasyOCR: pip-installierbar, Modelle kommen von GitHub
    (JaidedAI) statt von modelscope.cn -- das RapidOCR-CDN ist in vielen
    Firmen-/Heimnetzen blockiert und hinterlaesst dann kaputte Modelldateien.
    ``languages`` ist eine Kommaliste ("de,en"; Tesseract nutzt eigene Codes
    wie "deu,eng").
    """
    langs = [part.strip() for part in languages.split(",") if part.strip()]
    if engine == "easyocr":
        from docling.datamodel.pipeline_options import EasyOcrOptions

        return EasyOcrOptions(lang=langs or ["de", "en"])
    if engine == "tesseract":
        from docling.datamodel.pipeline_options import TesseractCliOcrOptions

        return TesseractCliOcrOptions(lang=langs or ["deu", "eng"])
    if engine == "rapidocr":
        from docling.datamodel.pipeline_options import RapidOcrOptions

        # torch-Backend: onnxruntime ist nicht Teil der Installation.
        return RapidOcrOptions(backend="torch")
    raise ValueError(
        f"Unbekannte OCR-Engine {engine!r} (erlaubt: easyocr, tesseract, rapidocr)"
    )


def check_ocr_engine(config: ConverterConfig) -> str | None:
    """Vorab-Pruefung der OCR-Engine; Warntext oder None.

    Wichtigster Fall: Tesseract ist als CLI-Aufruf realisiert -- fehlt die
    Installation, scheitert sonst JEDE Datei des Laufs einzeln mit derselben
    Meldung (real passiert: 3000+ Fehler statt einem klaren Hinweis vorab).
    """
    if not config.do_ocr:
        return None
    if config.ocr_engine == "tesseract" and shutil.which("tesseract") is None:
        return (
            "Die gewählte OCR-Engine Tesseract ist auf diesem Rechner nicht "
            "installiert (Befehl „tesseract“ nicht gefunden). Entweder die "
            "Standard-Engine EasyOCR wählen oder Tesseract installieren "
            "(Windows: UB-Mannheim-Installer, Sprache „German“ mitwählen) "
            "und neu starten."
        )
    return None


def check_paths(
    input_dir: os.PathLike | str | None,
    output_dir: os.PathLike | str | None,
) -> str | None:
    """Erkennt Quell-/Ziel-Konstellationen, die zwangslaeufig zu "0 Dateien
    gefunden" fuehren; Fehlertext oder None.

    Der Zielordner wird beim Scan bewusst ausgeschlossen, damit erzeugte
    Markdown-Dateien nicht beim naechsten Lauf selbst als Quelle gelten.
    Ist das Ziel identisch mit der Quelle (oder liegt die Quelle im Ziel),
    schliesst das ALLES aus -- das muss als klare Meldung sichtbar sein
    statt als stilles "0 unterstuetzte Dateien".

    Erlaubt bleibt der umgekehrte Fall (Ziel als Unterordner der Quelle):
    dann wird nur dieser Teilbereich ausgeschlossen.
    """
    if not input_dir or not output_dir:
        return None
    src = Path(input_dir).resolve()
    dst = Path(output_dir).resolve()
    if src == dst:
        return (
            "Quell- und Ziel-Ordner sind identisch. Der Ziel-Vault-Ordner "
            "wird beim Scan ausgeschlossen, damit erzeugte Markdown-Dateien "
            "nicht erneut als Quelle gelten – so würden 0 Dateien gefunden. "
            "Bitte einen eigenen Zielordner wählen, z. B. einen "
            "Nachbarordner („Hunter Vault“) oder einen Unterordner des "
            "Quellordners („…\\Vault“)."
        )
    if src.is_relative_to(dst):
        return (
            "Der Quellordner liegt innerhalb des Ziel-Vault-Ordners – damit "
            "würde der gesamte Quellbereich beim Scan ausgeschlossen "
            "(0 Dateien). Bitte das Ziel außerhalb des Quellordners oder "
            "als Unterordner des Quellordners wählen."
        )
    return None


def build_converter(
    config: ConverterConfig | None = None,
    pdf_backend: str | None = None,
):
    """Erzeugt einen konfigurierten Docling ``DocumentConverter``.

    Der Import von Docling passiert bewusst lazy (erst hier), damit das Modul
    auch ohne installiertes Docling importierbar bleibt -- z. B. fuer
    ``discover_files`` oder die Unit-Tests der Pfadlogik.

    ``pdf_backend="pypdfium"`` verwendet den alternativen pypdfium-Parser
    statt docling-parse -- Fallback fuer PDFs, die docling-parse mit
    "Inconsistent number of pages"/"not valid" ablehnt.
    """
    if config is None:
        config = ConverterConfig()

    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption

    pipeline_options = PdfPipelineOptions()
    pipeline_options.do_ocr = config.do_ocr
    pipeline_options.do_table_structure = config.do_table_structure
    pipeline_options.generate_picture_images = config.generate_picture_images
    pipeline_options.images_scale = config.images_scale
    if config.do_ocr:
        pipeline_options.ocr_options = _make_ocr_options(
            config.ocr_engine, config.ocr_languages
        )
    if config.do_table_structure:
        # Zellen-Matching verbessert die Tabellenrekonstruktion.
        pipeline_options.table_structure_options.do_cell_matching = True

    pdf_option_kwargs: dict = {"pipeline_options": pipeline_options}
    if pdf_backend == "pypdfium":
        from docling.backend.pypdfium2_backend import PyPdfiumDocumentBackend

        pdf_option_kwargs["backend"] = PyPdfiumDocumentBackend

    # Bilder (PNG/JPG/TIFF/...) routet Docling durch die PDF-Pipeline --
    # nur mit expliziter Option greifen OCR-/Bild-Einstellungen auch dort.
    from docling.document_converter import ImageFormatOption

    return DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(**pdf_option_kwargs),
            InputFormat.IMAGE: ImageFormatOption(pipeline_options=pipeline_options),
        }
    )


def discover_files(
    input_dir: os.PathLike | str,
    extensions: Iterable[str] = SUPPORTED_EXTENSIONS,
    exclude_dirs: Iterable[os.PathLike | str | None] = (),
) -> list[Path]:
    """Findet rekursiv alle unterstuetzten Dateien unterhalb von ``input_dir``.

    Ergebnis ist stabil sortiert (deterministische Reihenfolge/ETA). Versteckte
    Verzeichnisse (``.git`` etc.) und typische temporaere Office-Sperrdateien
    (``~$...``) werden uebersprungen.

    ``exclude_dirs`` blendet Verzeichnisse aus. Wichtigster Fall: liegt der
    Zielordner (Vault) innerhalb des Quellordners, wuerden die erzeugten
    ``.md``-Dateien beim naechsten Lauf selbst als Quelle erkannt -- Aufrufer
    uebergeben deshalb Ziel- und Archivordner.
    """
    input_path = Path(input_dir)
    exts = {e.lower() for e in extensions}
    excluded: list[Path] = []
    for e in exclude_dirs:
        if e:
            # resolve() statt absolute(): normalisiert Symlinks und
            # ../-Schreibweisen, sonst laesst sich der Ausschluss aushebeln.
            excluded.append(Path(e).resolve())
    files: list[Path] = []
    for path in input_path.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in exts:
            continue
        if path.name.startswith("~$"):
            continue
        if any(part.startswith(".") for part in path.relative_to(input_path).parts[:-1]):
            continue
        if excluded:
            p_abs = path.resolve()
            if any(ex == p_abs or ex in p_abs.parents for ex in excluded):
                continue
        files.append(path)
    return sorted(files)


# ---------------------------------------------------------------------------
# Zielordner-/Vault-Analyse und Integrationsplan
# ---------------------------------------------------------------------------
# Damit das Tool fuer beliebige Nutzer und Zielordner funktioniert: der
# Zielordner wird analysiert (Obsidian-Vault? Logseq? leer? bestehende
# Struktur?), daraus ein sinnvoller Integrationsplan abgeleitet und dieser
# EINMAL fuer den gesamten Batch zur Bestaetigung angezeigt.

# Verzeichnisse, die bei der Analyse nicht als inhaltliche Ordner zaehlen.
_VAULT_SKIP_DIRS = {".obsidian", ".trash", ".git", ".stfolder", ".stversions",
                    "logseq", ".DS_Store"}
# Uebliche Namen fuer Anhang-/Attachment-Ordner in bestehenden Vaults.
_ATTACHMENT_FOLDER_NAMES = ("attachments", "assets", "_attachments", "_media",
                            "media", "files", "resources")


@dataclass
class VaultProfile:
    """Ergebnis der Zielordner-Analyse."""

    target_path: str
    exists: bool
    is_empty: bool
    vault_type: str            # "obsidian" | "logseq" | "folder" | "new"
    note_count: int
    top_level_folders: list[str]
    attachment_folder_raw: str | None      # roher Wert aus .obsidian/app.json
    attachment_folder_resolved: str | None  # Ordnername oder None (=Wurzel)
    attachment_note_relative: bool             # Anhaenge neben der Notiz?
    uses_frontmatter: bool | None
    observations: list[str]


def _read_obsidian_attachment_setting(target: Path) -> str | None:
    """Liest ``attachmentFolderPath`` aus ``.obsidian/app.json`` (falls vorhanden)."""
    app_json = target / ".obsidian" / "app.json"
    if not app_json.is_file():
        return None
    try:
        data = json.loads(app_json.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    val = data.get("attachmentFolderPath")
    return val if isinstance(val, str) else None


def _detect_frontmatter_usage(target: Path, sample_limit: int = 25) -> bool | None:
    """Prueft an einer Stichprobe, ob bestehende Notizen YAML-Frontmatter nutzen."""
    checked = 0
    with_fm = 0
    for root, dirs, filenames in os.walk(target):
        dirs[:] = [d for d in dirs if d not in _VAULT_SKIP_DIRS and not d.startswith(".")]
        for name in filenames:
            if not name.lower().endswith(".md"):
                continue
            try:
                with open(Path(root) / name, encoding="utf-8", errors="ignore") as fh:
                    first = fh.readline().strip()
            except OSError:
                continue
            checked += 1
            if first == "---":
                with_fm += 1
            if checked >= sample_limit:
                break
        if checked >= sample_limit:
            break
    if checked == 0:
        return None
    return with_fm >= max(1, checked // 2)


def analyze_vault(target_dir: os.PathLike | str) -> VaultProfile:
    """Analysiert den Zielordner und liefert ein ``VaultProfile``.

    Erkennt Obsidian-Vaults (``.obsidian/``), Logseq-Graphen (``logseq/``),
    bestehende Nicht-Vault-Ordner und noch nicht existierende Ziele. Liest
    dabei -- soweit moeglich -- die Anhang-Konvention und die Frontmatter-
    Nutzung, um die spaetere Einordnung an den Vault anzupassen.
    """
    target = Path(target_dir)
    observations: list[str] = []

    if not target.exists():
        observations.append("Ziel existiert noch nicht – wird als neuer Ordner angelegt.")
        return VaultProfile(
            target_path=str(target), exists=False, is_empty=True, vault_type="new",
            note_count=0, top_level_folders=[], attachment_folder_raw=None,
            attachment_folder_resolved=None, attachment_note_relative=False,
            uses_frontmatter=None, observations=observations,
        )

    is_obsidian = (target / ".obsidian").is_dir()
    is_logseq = (target / "logseq").is_dir()

    # Inhaltliche Top-Level-Ordner (ohne Konfig-/Versteckt-Ordner).
    top_folders = sorted(
        p.name for p in target.iterdir()
        if p.is_dir() and p.name not in _VAULT_SKIP_DIRS and not p.name.startswith(".")
    )

    # Anzahl vorhandener Notizen (begrenzt gezaehlt, um grosse Vaults zu schonen).
    note_count = 0
    for _root, dirs, filenames in os.walk(target):
        dirs[:] = [d for d in dirs if d not in _VAULT_SKIP_DIRS and not d.startswith(".")]
        note_count += sum(1 for n in filenames if n.lower().endswith(".md"))
        if note_count > 100000:
            break

    is_empty = note_count == 0 and not top_folders

    # Anhang-Konvention bestimmen.
    attach_raw = _read_obsidian_attachment_setting(target)
    attach_note_relative = bool(attach_raw and attach_raw.startswith("./"))
    if attach_raw in (None, "", "/"):
        attach_resolved: str | None = None
    elif attach_note_relative:
        attach_resolved = attach_raw[2:].strip("/") or None
    else:
        attach_resolved = attach_raw.strip("/") or None

    # Falls Obsidian keine explizite Angabe hat: bestehenden Anhang-Ordner suchen.
    if attach_resolved is None and not attach_note_relative:
        for cand in _ATTACHMENT_FOLDER_NAMES:
            if any(f.lower() == cand for f in top_folders):
                match = next(f for f in top_folders if f.lower() == cand)
                attach_resolved = match
                observations.append(f"Bestehender Anhang-Ordner erkannt: „{match}“.")
                break

    if is_obsidian:
        vault_type = "obsidian"
        observations.append("Obsidian-Vault erkannt (.obsidian/).")
        if attach_raw is not None:
            where = ("neben der jeweiligen Notiz" if attach_note_relative
                     else (f"„{attach_resolved}“" if attach_resolved else "im Vault-Stamm"))
            observations.append(f"Obsidian legt Anhänge {where} ab.")
    elif is_logseq:
        vault_type = "logseq"
        observations.append("Logseq-Graph erkannt (logseq/). Notizen → pages/, Anhänge → assets/.")
    else:
        vault_type = "folder"
        if is_empty:
            observations.append("Leerer Ordner – Struktur wird neu aufgebaut.")
        else:
            observations.append("Bestehender Ordner (kein Vault-Marker gefunden).")

    uses_fm = _detect_frontmatter_usage(target) if note_count else None
    if uses_fm is True:
        observations.append("Bestehende Notizen nutzen Frontmatter/Properties.")
    elif uses_fm is False:
        observations.append("Bestehende Notizen nutzen kein Frontmatter.")

    return VaultProfile(
        target_path=str(target), exists=True, is_empty=is_empty, vault_type=vault_type,
        note_count=note_count, top_level_folders=top_folders,
        attachment_folder_raw=attach_raw, attachment_folder_resolved=attach_resolved,
        attachment_note_relative=attach_note_relative, uses_frontmatter=uses_fm,
        observations=observations,
    )


# Standard-Unterordner fuer Importe in einen bestehenden Vault.
DEFAULT_IMPORT_SUBDIR = "Docling Import"


def recommend_config(
    profile: VaultProfile, base: ConverterConfig | None = None
) -> ConverterConfig:
    """Leitet aus einem ``VaultProfile`` einen empfohlenen Integrationsplan ab.

    Grundhaltung: einen bestehenden, kuratierten Vault nicht zumuellen — Importe
    kommen in einen eigenen Unterordner, folgen aber der Anhang- und
    Frontmatter-Konvention des Vaults. Leere/neue Ziele werden frei aufgebaut.
    """
    cfg = base or ConverterConfig()
    cfg.mirror_structure = True

    if profile.vault_type == "obsidian":
        cfg.notes_subdir = "" if profile.is_empty else DEFAULT_IMPORT_SUBDIR
        if profile.attachment_note_relative:
            cfg.attachments_mode = "adjacent"
        else:
            cfg.attachments_mode = "central"
            cfg.attachments_subdir = profile.attachment_folder_resolved or "assets"
        cfg.add_frontmatter = profile.uses_frontmatter is not False
    elif profile.vault_type == "logseq":
        cfg.notes_subdir = "pages"
        cfg.attachments_mode = "central"
        cfg.attachments_subdir = "assets"
        cfg.add_frontmatter = True
    else:  # "folder" | "new"
        cfg.notes_subdir = "" if profile.is_empty else DEFAULT_IMPORT_SUBDIR
        cfg.attachments_mode = "central"
        cfg.attachments_subdir = profile.attachment_folder_resolved or "assets"
        cfg.add_frontmatter = True

    return cfg


def describe_plan(profile: VaultProfile, config: ConverterConfig) -> list[str]:
    """Menschlich lesbare Zusammenfassung des Integrationsplans (fuer Bestaetigung)."""
    target = profile.target_path
    notes_loc = f"{target}/{config.notes_subdir}".rstrip("/") if config.notes_subdir else target
    struct = "gespiegelte Quellstruktur" if config.mirror_structure else "flach (ohne Unterordner)"
    if config.attachments_mode == "adjacent":
        attach = "neben der jeweiligen Notiz (per-Notiz-Ordner)"
    else:
        attach = f"zentral in „{config.attachments_subdir}/“"
    lines = [
        f"Ziel: {target}",
        f"Notizen (.md) → {notes_loc}  ({struct})",
        f"Anhänge/Bilder → {attach}",
        f"Frontmatter/Properties: {'ja' if config.add_frontmatter else 'nein'}",
    ]
    return lines


def _asset_key(rel_path: Path) -> str:
    """Kollisionsfreier Ordnername fuer die Assets einer Quelldatei.

    Aus ``berichte/2024/q1.pdf`` wird ``berichte__2024__q1`` -- so kollidieren
    gleichnamige Dateien aus verschiedenen Unterordnern nicht im ``assets/``-
    Verzeichnis.
    """
    stem_rel = rel_path.with_suffix("")
    return "__".join(stem_rel.parts)


def _yaml_frontmatter(fields: dict[str, object]) -> str:
    """Baut einen YAML-Frontmatter-Block. Strings werden JSON-quoted.

    JSON-Doppelquoting liefert einen gueltigen YAML-Skalar und escaped
    Sonderzeichen (Backslashes in Windows-Pfaden, Doppelpunkte, ...) korrekt.
    """
    lines = ["---"]
    for key, value in fields.items():
        if value is None:
            continue
        if isinstance(value, bool):
            lines.append(f"{key}: {str(value).lower()}")
        elif isinstance(value, (int, float)):
            lines.append(f"{key}: {value}")
        else:
            lines.append(f"{key}: {json.dumps(str(value), ensure_ascii=False)}")
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


# Klassifizierung der haeufigsten Fehlerursachen -> (Kategorie, Klartext-Hinweis).
# Reihenfolge = Prioritaet; erste passende Regel gewinnt. Gematcht wird gegen
# Exceptionklasse + Nachricht + Traceback (kleingeschrieben).
_ERROR_RULES: list[tuple[tuple[str, ...], str, str]] = [
    (
        ("unexpected eof", "recall_on_data_access"),
        "cloud-platzhalter",
        "Datei liegt vermutlich nur als Cloud-Platzhalter vor (OneDrive "
        "„Dateien bei Bedarf“) und ist lokal unvollständig. Ordner in "
        "OneDrive auf „Immer auf diesem Gerät behalten“ stellen und erneut "
        "ausführen.",
    ),
    (
        ("storage has wrong byte size", "pickle data was truncated",
         "file exists but is invalid", "modelscope", "rapidocr"),
        "ocr-modelle",
        "OCR-Modelldateien fehlen oder sind beschädigt – der RapidOCR-Download "
        "(modelscope.cn) ist oft blockiert. Auf die Standard-Engine EasyOCR "
        "wechseln (Modelle von GitHub) oder den RapidOCR-Modellordner löschen "
        "und mit funktionierendem Netzzugang erneut laden.",
    ),
    (
        ("downloading detection model", "downloading recognition model",
         "easyocr"),
        "ocr-modelle",
        "Die EasyOCR-Modelle konnten nicht heruntergeladen werden (GitHub "
        "blockiert oder offline). Entweder Netzzugang herstellen, die "
        "Modelldateien manuell in den EasyOCR-Modellordner legen "
        "(~/.EasyOCR/model) oder auf Tesseract wechseln – der arbeitet "
        "komplett lokal ohne Downloads.",
    ),
    (
        ("tesseract is not available", "tesseract_cmd", "tesserocr"),
        "ocr-engine",
        "Die gewählte OCR-Engine Tesseract ist auf diesem Rechner nicht "
        "installiert oder nicht im PATH. Entweder in der Seitenleiste auf "
        "die Standard-Engine EasyOCR wechseln oder Tesseract installieren "
        "(Windows: UB-Mannheim-Installer, Sprache „German“ mitwählen) und "
        "das Dashboard neu starten.",
    ),
    (
        ("inconsistent number of pages", "conversion failed for"),
        "pdf-parser",
        "Der Standard-PDF-Parser (docling-parse) konnte das Dokument nicht "
        "laden. doc2vault versucht solche PDFs automatisch erneut mit dem "
        "alternativen pypdfium-Parser – schlägt auch das fehl, ist die Datei "
        "vermutlich beschädigt oder verwendet ein exotisches PDF-Format.",
    ),
    (
        ("terminated abruptly", "brokenprocesspool", "prozess abgestürzt"),
        "prozessabsturz",
        "Ein Konvertierungsprozess ist abgestürzt – meist Speicher bei sehr "
        "großen/komplexen PDFs. Parallele Prozesse und Bildauflösung "
        "reduzieren; die Datei einzeln erneut versuchen.",
    ),
    (
        ("password", "encrypted", "passwort", "verschlüss", "decrypt", "is protected"),
        "passwortgeschützt",
        "Datei ist passwortgeschützt oder verschlüsselt – vor der Konvertierung entsperren.",
    ),
    (
        ("no such file", "does not exist", "filenotfound", "permission denied"),
        "nicht lesbar",
        "Datei nicht gefunden oder keine Leserechte.",
    ),
    (
        ("memoryerror", "cannot allocate", "out of memory", "oom", "killed",
         "bad_alloc", "not enough memory", "unable to allocate",
         "defaultcpuallocator", "teilkonvertierung"),
        "speicher",
        "Zu wenig Arbeitsspeicher – Anzahl paralleler Prozesse reduzieren.",
    ),
    (
        ("timeout", "timed out"),
        "timeout",
        "Zeitüberschreitung bei der Verarbeitung – Datei ist möglicherweise sehr groß oder komplex.",
    ),
    (
        ("unsupported", "no backend", "not supported", "unknown format"),
        "nicht unterstützt",
        "Format/Variante wird von Docling nicht unterstützt.",
    ),
    (
        ("corrupt", "damaged", "eof marker", "not a pdf", "invalid", "broken",
         "cannot read", "failed to parse", "malformed", "bad", "truncated",
         "zip file", "not a zip"),
        "beschädigt",
        "Datei ist vermutlich beschädigt oder kein gültiges Dokument.",
    ),
]


def _classify_error(text: str) -> tuple[str, str]:
    """Ordnet einen Fehlertext einer Kategorie + Klartext-Hinweis zu."""
    low = text.lower()
    for keywords, category, hint in _ERROR_RULES:
        if any(k in low for k in keywords):
            return category, hint
    return "fehler", "Unerwarteter Fehler – vollständige Ursache siehe Traceback."


def xlsx_sheet_names(path: os.PathLike | str) -> list[str]:
    """Liest die Blattnamen einer XLSX-Datei ohne Zusatzabhaengigkeiten.

    XLSX ist ein ZIP-Archiv; die Blaetter stehen in ``xl/workbook.xml``.
    Bei beschaedigten/untypischen Dateien wird eine leere Liste geliefert --
    die eigentliche Fehlerbehandlung uebernimmt dann die Konvertierung.
    """
    import zipfile
    from xml.etree import ElementTree

    try:
        with zipfile.ZipFile(path) as zf:
            with zf.open("xl/workbook.xml") as fh:
                root = ElementTree.parse(fh).getroot()
    except Exception:  # noqa: BLE001 -- defekte Datei o. ae.
        return []
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [el.get("name", "") for el in root.findall("m:sheets/m:sheet", ns)]


def _stem_collides(source: Path) -> bool:
    """True, wenn im selben Ordner eine WEITERE unterstuetzte Datei mit
    gleichem Stamm liegt (z. B. Word-Datei + daraus exportiertes PDF)."""
    try:
        stem = source.stem.lower()
        for p in source.parent.iterdir():
            if (
                p != source
                and p.suffix.lower() in SUPPORTED_EXTENSIONS
                and p.stem.lower() == stem
                and p.is_file()
            ):
                return True
    except OSError:
        return False
    return False


def _trim_xlsx(source: Path, keep: int) -> Path:
    """Erzeugt eine temporaere Kopie der Arbeitsmappe mit den ersten ``keep``
    Blaettern. openpyxl ist eine Docling-Abhaengigkeit und daher zur Laufzeit
    verfuegbar; der Import bleibt lazy, damit das Modul ohne sie importierbar
    ist."""
    import tempfile

    from openpyxl import load_workbook

    workbook = load_workbook(source)
    for name in workbook.sheetnames[keep:]:
        del workbook[name]
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix=f"{source.stem}_trimmed_")
    os.close(fd)
    try:
        workbook.save(tmp)
    except BaseException:
        # save fehlgeschlagen (Platte voll, exotische Features): Temp-Datei
        # nicht liegen lassen -- das finally des Aufrufers kennt sie noch nicht.
        os.unlink(tmp)
        raise
    return Path(tmp)


def _apply_post_action(
    source: Path,
    config: ConverterConfig,
    input_root: os.PathLike | str | None,
) -> str | None:
    """Verschiebt/loescht das Original nach erfolgreicher Konvertierung.

    Gibt das Zielverzeichnis (bei ``archive``) bzw. ``"<geloescht>"`` zurueck,
    oder ``None`` wenn nichts getan wurde.
    """
    if config.on_success == "archive" and config.archive_dir:
        try:
            rel = source.relative_to(input_root) if input_root else Path(source.name)
        except ValueError:
            rel = Path(source.name)
        dest = Path(config.archive_dir) / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(source), str(dest))
        return str(dest)
    if config.on_success == "delete":
        source.unlink()
        return "<geloescht>"
    return None


def convert_single_file(
    source_path: os.PathLike | str,
    output_dir: os.PathLike | str,
    input_root: os.PathLike | str | None = None,
    config: ConverterConfig | None = None,
    converter=None,
) -> ConversionResult:
    """Konvertiert eine Datei nach Markdown und extrahiert eingebettete Bilder.

    Die Verzeichnisstruktur relativ zu ``input_root`` wird im ``output_dir``
    (dem Vault) gespiegelt. Bilder landen unter ``output_dir/assets/<key>/``.
    Fehler werden abgefangen, klassifiziert und mit vollem Traceback als
    ``ConversionResult(success=False)`` zurueckgegeben -- ein kaputtes Dokument
    bricht den Batch nicht ab.
    """
    if config is None:
        config = ConverterConfig()
    start = time.perf_counter()
    source = Path(source_path)
    out_root = Path(output_dir)

    try:
        rel = source.relative_to(input_root) if input_root else Path(source.name)
    except ValueError:
        # Quelle liegt nicht unter input_root -> flach ablegen.
        rel = Path(source.name)

    # Report.pdf + Report.docx im selben Ordner wuerden beide auf Report.md
    # (und denselben Asset-Ordner) abgebildet und sich still ueberschreiben --
    # bei Kollision die Endung in den Namen aufnehmen (deterministisch, auch
    # ueber parallele Worker hinweg).
    if _stem_collides(source):
        tag = source.suffix.lstrip(".").lower()
        rel = rel.with_name(f"{rel.stem} ({tag}){rel.suffix}")

    # Notiz-Ablage: optional in einen Unterordner (Vault-Einordnung) und
    # optional mit gespiegelter Quellstruktur.
    note_rel = rel if config.mirror_structure else Path(rel.name)
    notes_base = out_root / config.notes_subdir if config.notes_subdir else out_root
    md_path = notes_base / note_rel.with_suffix(".md")

    # Anhang-Ablage: zentral (ein Ordner nach Vault-Konvention) oder neben der
    # Notiz (per-Notiz-Ordner). _asset_key vermeidet Namenskollisionen.
    if config.attachments_mode == "adjacent":
        assets_dir = md_path.parent / f"{md_path.stem}_assets"
    else:
        assets_dir = out_root / (config.attachments_subdir or "assets") / _asset_key(rel)

    # XLSX-Sonderfall: Arbeitsmappen mit zu vielen Blaettern begrenzen oder
    # ueberspringen (Zaehlung ist billig, kein Docling noetig).
    convert_input: Path = source
    trimmed_tmp: Path | None = None
    sheets_total: int | None = None
    sheets_converted: int | None = None
    if source.suffix.lower() == ".xlsx" and config.xlsx_sheet_limit > 0:
        names = xlsx_sheet_names(source)
        if len(names) > config.xlsx_sheet_limit:
            if config.xlsx_on_limit == "skip":
                return ConversionResult(
                    source_path=str(source),
                    success=False,
                    duration_s=time.perf_counter() - start,
                    error=f"{len(names)} Blätter, Limit {config.xlsx_sheet_limit}",
                    error_category="zu viele sheets",
                    error_hint="Arbeitsmappe übersprungen. Limit erhöhen, den "
                    "Modus „nur erste Blätter“ wählen oder die Datei gezielt "
                    "einzeln konvertieren.",
                )
            sheets_total = len(names)
            sheets_converted = config.xlsx_sheet_limit

    try:
        from docling_core.types.doc import ImageRefMode

        # Datei einmal vollstaendig sequenziell lesen: zwingt Cloud-Platzhalter
        # (OneDrive "Dateien bei Bedarf") zum Herunterladen, bevor Docling mit
        # partiellen Reads auf unvollstaendigen Daten scheitert.
        with open(source, "rb") as fh:
            while fh.read(1 << 20):
                pass

        if converter is None:
            converter = build_converter(config)

        if sheets_total is not None:
            trimmed_tmp = _trim_xlsx(source, config.xlsx_sheet_limit)
            convert_input = trimmed_tmp

        result = converter.convert(convert_input)

        # Teilkonvertierung ist KEIN Erfolg: bei Speicherdruck scheitern
        # einzelne Seiten in Docling ("Stage preprocess failed ...
        # std::bad_alloc"), der Rest laeuft weiter -- ohne diese Pruefung
        # entstuende eine Notiz mit still fehlenden Seiten. Als Fehler
        # gemeldet greift stattdessen der automatische reduzierte
        # Zweitversuch im isolierten Einzelprozess.
        status = str(getattr(result, "status", "")).lower()
        if "partial" in status:
            errors = getattr(result, "errors", None) or []
            detail = "; ".join(
                str(getattr(e, "error_message", e)) for e in errors[:10]
            )
            raise RuntimeError(
                f"Teilkonvertierung: {len(errors) or 'einige'} Seite(n)/"
                f"Stufe(n) fehlgeschlagen – Notiz wäre unvollständig. "
                f"{detail[:400]}"
            )

        doc = result.document

        md_path.parent.mkdir(parents=True, exist_ok=True)

        # save_as_markdown schreibt die .md samt referenzierter Bilder in
        # assets_dir und setzt relative Links von der .md dorthin.
        doc.save_as_markdown(
            md_path,
            artifacts_dir=assets_dir,
            image_mode=ImageRefMode.REFERENCED,
        )

        num_images = 0
        assets_rel: str | None = None
        if assets_dir.exists():
            num_images = sum(1 for p in assets_dir.iterdir() if p.is_file())
            if num_images:
                assets_rel = os.path.relpath(assets_dir, out_root).replace(os.sep, "/")

        body = md_path.read_text(encoding="utf-8")

        # Docling referenziert Bilder u. U. mit absolutem Pfad, wenn der
        # Asset-Ordner nicht unterhalb des Notiz-Ordners liegt (zentrale
        # Ablage + verschachtelte Notiz). Fuer Obsidian muessen die Links
        # relativ zur Notiz sein.
        if num_images:
            abs_prefix = assets_dir.absolute().as_posix()
            if abs_prefix in body:
                rel_prefix = Path(
                    os.path.relpath(assets_dir, md_path.parent)
                ).as_posix()
                body = body.replace(abs_prefix, rel_prefix)

        if config.add_frontmatter:
            body = _yaml_frontmatter(
                {
                    "source": source.name,
                    "original_path": str(source.resolve()),
                    "assets_folder": assets_rel,
                    "converted_at": datetime.now(timezone.utc)
                    .isoformat(timespec="seconds"),
                    "converter": "docling",
                    "sheets_total": sheets_total,
                    "sheets_converted": sheets_converted,
                }
            ) + body

        md_path.write_text(body, encoding="utf-8")

        # Original erst NACH erfolgreichem Schreiben archivieren/loeschen.
        # Ein Fehler HIER (Datei in Word geoeffnet -> PermissionError) macht
        # die Konvertierung nicht ungueltig -- die .md ist fertig; nur die
        # Nachaktion wird als Warnung vermerkt statt als Fehlschlag.
        moved_to = None
        post_action_error = None
        try:
            moved_to = _apply_post_action(source, config, input_root)
        except Exception as exc:  # noqa: BLE001
            post_action_error = f"{type(exc).__name__}: {exc}"

        return ConversionResult(
            source_path=str(source),
            success=True,
            output_path=str(md_path),
            assets_folder=str(assets_dir) if assets_rel else None,
            num_images=num_images,
            duration_s=time.perf_counter() - start,
            moved_to=moved_to,
            post_action_error=post_action_error,
        )
    except Exception as exc:  # noqa: BLE001 -- Batch soll robust weiterlaufen
        detail = traceback.format_exc()
        concise = f"{type(exc).__name__}: {exc}".strip()
        category, hint = _classify_error(f"{type(exc).__name__} {exc}\n{detail}")
        return ConversionResult(
            source_path=str(source),
            success=False,
            duration_s=time.perf_counter() - start,
            error=concise,
            error_category=category,
            error_hint=hint,
            error_detail=detail.strip(),
        )
    finally:
        if trimmed_tmp is not None:
            trimmed_tmp.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Prozess-Pool-Helfer
# ---------------------------------------------------------------------------
# Docling ist CPU-lastig, deshalb ProcessPoolExecutor statt Threads (der GIL
# wuerde bei Threads bremsen). Jeder Worker-Prozess baut EINEN Converter beim
# Start (init_worker) und verwendet ihn fuer alle ihm zugewiesenen Dateien --
# der teure Modell-/Pipeline-Aufbau passiert so nur einmal pro Prozess.

_WORKER_CONVERTER = None
_WORKER_CONVERTER_REDUCED = None
# pypdfium-Fallback getrennt fuer Voll- und reduzierte Konfiguration: ein
# gemeinsamer Cache wuerde je nach erstem Ausloeser entweder Bilder still
# weglassen oder Riesenseiten wieder mit voller Skalierung rendern.
_WORKER_CONVERTER_PDFIUM = None
_WORKER_CONVERTER_PDFIUM_REDUCED = None
_WORKER_CONFIG: ConverterConfig | None = None
_WORKER_OUTPUT: Path | None = None
_WORKER_ROOT: Path | None = None


def init_worker(config: ConverterConfig, output_dir: str, input_root: str) -> None:
    """Initialisiert einen Worker-Prozess (baut den Converter einmalig)."""
    global _WORKER_CONVERTER, _WORKER_CONVERTER_REDUCED
    global _WORKER_CONVERTER_PDFIUM, _WORKER_CONVERTER_PDFIUM_REDUCED
    global _WORKER_CONFIG, _WORKER_OUTPUT, _WORKER_ROOT
    _mute_streamlit_bare_mode_warning()
    _mute_torch_pin_memory_warning()
    _mute_worker_progress_bars()
    _WORKER_CONFIG = config
    _WORKER_OUTPUT = Path(output_dir)
    _WORKER_ROOT = Path(input_root)
    _WORKER_CONVERTER = build_converter(config)
    # Alle Fallback-Converter entstehen lazy, nur wenn sie gebraucht werden.
    _WORKER_CONVERTER_REDUCED = None
    _WORKER_CONVERTER_PDFIUM = None
    _WORKER_CONVERTER_PDFIUM_REDUCED = None


def convert_file_task(source_path: str) -> ConversionResult:
    """Pool-Task: konvertiert eine Datei mit dem Prozess-lokalen Converter.

    PDFs mit riesigen Seiten (CAD-Zeichnungen u. ae.) werden automatisch mit
    speicherschonenden Einstellungen konvertiert, statt den Worker mit
    ``std::bad_alloc`` zu sprengen.
    """
    global _WORKER_CONVERTER_REDUCED

    config = _WORKER_CONFIG
    converter = _WORKER_CONVERTER
    reduced = False
    if (
        config is not None
        and not _is_reduced(config)
        and source_path.lower().endswith(".pdf")
        and has_huge_pages(source_path)
    ):
        if _WORKER_CONVERTER_REDUCED is None:
            _WORKER_CONVERTER_REDUCED = build_converter(_reduced_config(config))
        converter = _WORKER_CONVERTER_REDUCED
        config = _reduced_config(config)
        reduced = True

    result = convert_single_file(
        source_path,
        _WORKER_OUTPUT,
        input_root=_WORKER_ROOT,
        config=config,
        converter=converter,
    )
    result.reduced_mode = reduced or result.reduced_mode

    # Fallback-Parser: docling-parse lehnt manche real existierenden PDFs mit
    # "Inconsistent number of pages: N!=-1" / "Input document is not valid"
    # ab (generischer ConversionError "Conversion failed for: ..."). pypdfium
    # laedt dieselben Dateien meist problemlos -> einmaliger zweiter Versuch.
    global _WORKER_CONVERTER_PDFIUM, _WORKER_CONVERTER_PDFIUM_REDUCED
    if (
        not result.success
        and source_path.lower().endswith(".pdf")
        and "conversion failed for" in (result.error or "").lower()
    ):
        try:
            # Cache passend zur effektiven Config waehlen: `config` ist an
            # dieser Stelle bereits die ggf. reduzierte Variante.
            if _is_reduced(config):
                if _WORKER_CONVERTER_PDFIUM_REDUCED is None:
                    _WORKER_CONVERTER_PDFIUM_REDUCED = build_converter(
                        config, pdf_backend="pypdfium"
                    )
                fallback_converter = _WORKER_CONVERTER_PDFIUM_REDUCED
            else:
                if _WORKER_CONVERTER_PDFIUM is None:
                    _WORKER_CONVERTER_PDFIUM = build_converter(
                        config, pdf_backend="pypdfium"
                    )
                fallback_converter = _WORKER_CONVERTER_PDFIUM
            retry = convert_single_file(
                source_path,
                _WORKER_OUTPUT,
                input_root=_WORKER_ROOT,
                config=config,
                converter=fallback_converter,
            )
        except Exception:  # noqa: BLE001 -- urspruengliches Ergebnis behalten
            retry = None
        if retry is not None and retry.success:
            retry.reduced_mode = result.reduced_mode
            retry.pdf_backend = "pypdfium"
            return retry
    return result


# Wie oft eine Datei, die beim Verarbeiten einen Worker-Absturz miterlebt hat,
# erneut versucht wird, bevor sie endgueltig als fehlgeschlagen gilt.
_CRASH_RETRY_LIMIT = 2


def _abort_pool(pool) -> None:
    """Sofort-Stopp eines Pools: laufende Worker beenden statt auf sie zu
    warten. Noetig fuer den Abbrechen-Button im Dashboard und Strg+C in der
    CLI -- das normale ``shutdown(wait=True)`` wuerde sonst minutenlang auf
    angefangene Konvertierungen warten."""
    try:
        pool.shutdown(wait=False, cancel_futures=True)
    except Exception:  # noqa: BLE001
        pass
    procs = getattr(pool, "_processes", None) or {}
    for proc in list(procs.values()):
        try:
            proc.terminate()
        except Exception:  # noqa: BLE001
            pass


def run_conversion_batch(
    files: list,
    config: ConverterConfig,
    output_dir: os.PathLike | str,
    input_root: os.PathLike | str,
    max_workers: int,
    progress=None,
    heartbeat=None,
) -> list[ConversionResult]:
    """Konvertiert eine Dateiliste parallel und uebersteht Worker-Abstuerze.

    Stuerzt ein Worker-Prozess hart ab (z. B. ``std::bad_alloc`` bei einer
    riesigen CAD-Zeichnung), reisst ``ProcessPoolExecutor`` normalerweise ALLE
    noch offenen Futures mit -- ein einziges Problemdokument liess so ganze
    Batches scheitern. Dieser Runner faengt den Pool-Bruch ab, startet den
    Pool neu und verarbeitet die restlichen Dateien weiter. Dateien, die
    mehrfach (``_CRASH_RETRY_LIMIT``) waehrend eines Absturzes in Arbeit
    waren, werden als "prozessabsturz" markiert statt endlos wiederholt.

    ``progress(done, total, result)`` wird fuer jedes Ergebnis aufgerufen.
    ``heartbeat()`` wird ~sekuendlich aufgerufen, solange Worker beschaeftigt
    sind -- das Dashboard nutzt das als Punkt, an dem ein Abbrechen-Klick
    greift (Streamlit unterbricht Skripte nur an UI-Aufrufen). Eine aus
    ``progress``/``heartbeat`` aufsteigende BaseException (Streamlit-Rerun,
    KeyboardInterrupt) beendet die laufenden Worker sofort.
    """
    from concurrent.futures import FIRST_COMPLETED, ProcessPoolExecutor, wait
    from concurrent.futures.process import BrokenProcessPool

    remaining = [str(f) for f in files]
    total = len(remaining)
    crash_seen: dict[str, int] = {}
    results: list[ConversionResult] = []
    # Speicher-/Absturzfaelle bekommen am Ende automatisch einen zweiten
    # Versuch mit reduzierten Einstellungen in einem isolierten Einzelprozess.
    retry_reduced: list[tuple[str, ConversionResult]] = []
    done = 0

    def _emit(res: ConversionResult) -> None:
        nonlocal done
        done += 1
        results.append(res)
        if progress:
            progress(done, total, res)

    def _collect_failure(res: ConversionResult) -> None:
        if res.error_category in ("speicher", "prozessabsturz"):
            retry_reduced.append((res.source_path, res))
        else:
            _emit(res)

    while remaining:
        crashed: list[str] = []
        pool = ProcessPoolExecutor(
            max_workers=max_workers,
            initializer=init_worker,
            initargs=(config, str(output_dir), str(input_root)),
        )
        try:
            try:
                futures = {pool.submit(convert_file_task, f): f for f in remaining}
                pending = set(futures)
                while pending:
                    done_now, pending = wait(
                        pending, timeout=1.0, return_when=FIRST_COMPLETED
                    )
                    if not done_now:
                        if heartbeat:
                            heartbeat()
                        continue
                    for future in done_now:
                        src = futures[future]
                        try:
                            res = future.result()
                        except BrokenProcessPool:
                            crashed.append(src)
                            continue
                        except Exception as exc:  # noqa: BLE001 -- Batch weiterfuehren
                            detail = traceback.format_exc()
                            category, hint = _classify_error(f"{exc}\n{detail}")
                            _collect_failure(ConversionResult(
                                source_path=src, success=False,
                                error=f"Pool-Fehler: {exc}",
                                error_category=category, error_hint=hint,
                                error_detail=detail.strip(),
                            ))
                            continue
                        if res.success:
                            _emit(res)
                        else:
                            _collect_failure(res)
                pool.shutdown(wait=True)
            except BrokenProcessPool:
                # Bruch beim Einreihen/Shutdown: alle noch nicht gemeldeten
                # Dateien gelten als potenziell betroffen.
                pool.shutdown(wait=False, cancel_futures=True)
                handled = {r.source_path for r in results}
                handled |= {src for src, _ in retry_reduced}
                crashed += [f for f in remaining
                            if f not in crashed and f not in handled]
        except BaseException:
            # Abbruch von aussen (Abbrechen-Button/Streamlit-Rerun, Strg+C):
            # Worker sofort beenden, nicht auf angefangene Dateien warten.
            _abort_pool(pool)
            raise

        next_round: list[str] = []
        for src in crashed:
            crash_seen[src] = crash_seen.get(src, 0) + 1
            if crash_seen[src] >= _CRASH_RETRY_LIMIT:
                category, hint = _classify_error("prozess abgestürzt")
                retry_reduced.append((src, ConversionResult(
                    source_path=src, success=False,
                    error="Worker-Prozess ist beim Verarbeiten abgestürzt "
                    "(vermutlich Speicher).",
                    error_category=category, error_hint=hint,
                )))
            else:
                next_round.append(src)
        remaining = next_round

    # Zweiter Versuch mit reduzierten Einstellungen (Bildskalierung 1.0,
    # ohne Bildextraktion), sequenziell und isoliert -- maximaler Speicher
    # fuer die eine Problemdatei, ein erneuter Absturz bleibt eingedaemmt.
    reduced = _reduced_config(config)
    # Startet KEIN Worker (kaputte Installation, blockierter Modell-Download),
    # bricht jeder Pool sofort -- ohne Kappe wuerde fuer jede der ggf.
    # tausenden Dateien ein weiterer zum Scheitern verurteilter Pool
    # gestartet (auf Windows je mehrere Sekunden Spawn-Zeit).
    ok_any = any(r.success for r in results)
    broken_pools = 0
    for src, first_fail in retry_reduced:
        if broken_pools >= 3 and not ok_any:
            first_fail.error_hint = (
                (first_fail.error_hint or "").rstrip() +
                " Die Worker-Prozesse konnten wiederholt gar nicht erst "
                "starten – vermutlich ein Installations- oder Modellproblem "
                "(Konsolen-Log prüfen), keine Einzeldatei-Ursache."
            ).strip()
            _emit(first_fail)
            continue
        retry_result: ConversionResult | None = None
        pool_broke = False
        pool = ProcessPoolExecutor(
            max_workers=1,
            initializer=init_worker,
            initargs=(reduced, str(output_dir), str(input_root)),
        )
        try:
            try:
                future = pool.submit(convert_file_task, src)
                pending = {future}
                while pending:
                    _, pending = wait(
                        pending, timeout=1.0, return_when=FIRST_COMPLETED
                    )
                    if pending and heartbeat:
                        heartbeat()   # Abbrechen-Klick greift auch hier
                retry_result = future.result()
            except BrokenProcessPool:
                retry_result = None
                pool_broke = True
            except Exception:  # noqa: BLE001
                retry_result = None
            pool.shutdown(wait=True)
        except BaseException:
            _abort_pool(pool)
            raise
        broken_pools = broken_pools + 1 if pool_broke else 0
        if retry_result is not None and retry_result.success:
            retry_result.reduced_mode = True
            _emit(retry_result)
        else:
            first_fail.error_hint = (
                (first_fail.error_hint or "").rstrip() +
                " Der automatische Wiederholungsversuch mit reduzierten "
                "Einstellungen (Bildskalierung 1.0, ohne Bildextraktion) "
                "ist ebenfalls fehlgeschlagen."
            ).strip()
            _emit(first_fail)

    return results


# ---------------------------------------------------------------------------
# CLI (nackte Nutzung ohne Streamlit)
# ---------------------------------------------------------------------------

def _run_cli(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Docling-Batch-Konvertierung (PDF/DOCX/XLSX -> Markdown)."
    )
    parser.add_argument("--input", "-i", required=True, help="Quellordner")
    parser.add_argument("--output", "-o", required=True, help="Ziel-Vault-Ordner")
    parser.add_argument(
        "--workers",
        "-w",
        type=int,
        default=max(1, (os.cpu_count() or 2) - 1),
        help="Anzahl paralleler Prozesse (Default: CPUs-1)",
    )
    parser.add_argument(
        "--ocr",
        action="store_true",
        help="OCR aktivieren (langsam; nur fuer gescannte PDFs)",
    )
    parser.add_argument(
        "--ocr-engine",
        choices=["easyocr", "tesseract", "rapidocr"],
        default="easyocr",
        help="OCR-Engine (Default easyocr: Modelle von GitHub statt "
        "modelscope.cn; tesseract erfordert lokale Installation)",
    )
    parser.add_argument(
        "--ocr-langs",
        default="de,en",
        help="OCR-Sprachen als Kommaliste (easyocr: de,en; tesseract: deu,eng)",
    )
    parser.add_argument(
        "--no-images",
        action="store_true",
        help="Keine eingebetteten Bilder extrahieren (reine Textkonvertierung)",
    )
    parser.add_argument(
        "--images-scale",
        type=float,
        default=2.0,
        help="Skalierung der extrahierten Bilder (Default 2.0)",
    )
    parser.add_argument(
        "--no-tables",
        action="store_true",
        help="Tabellenstruktur-Erkennung deaktivieren (schneller)",
    )
    parser.add_argument(
        "--xlsx-sheet-limit",
        type=int,
        default=0,
        help="Max. Blaetter je XLSX-Arbeitsmappe (0 = alle)",
    )
    parser.add_argument(
        "--xlsx-on-limit",
        choices=["limit", "skip"],
        default="limit",
        help="Bei Ueberschreitung: limit = nur erste Blaetter, skip = Datei "
        "ueberspringen",
    )
    parser.add_argument(
        "--on-success",
        choices=["keep", "archive", "delete"],
        default="keep",
        help="Was mit erfolgreich konvertierten Originalen passiert "
        "(keep=behalten, archive=verschieben, delete=loeschen)",
    )
    parser.add_argument(
        "--archive-dir",
        default=None,
        help="Zielordner fuer --on-success archive",
    )
    parser.add_argument(
        "--notes-subdir",
        default=None,
        help="Unterordner im Ziel fuer die Notizen (ueberschreibt Empfehlung; "
        "\"\" = direkt in die Ziel-Wurzel)",
    )
    parser.add_argument(
        "--attachments-subdir",
        default=None,
        help="Name des zentralen Anhang-Ordners (ueberschreibt Empfehlung)",
    )
    parser.add_argument(
        "--no-frontmatter",
        action="store_true",
        help="Kein YAML-Frontmatter voranstellen",
    )
    parser.add_argument(
        "--build-vault",
        action="store_true",
        help="Nach der Konvertierung den Vault-Build ausfuehren: Notizen nach "
        "Inbox/, Bilder nach Attachments/<notiz>/, Wikilinks, normiertes "
        "Frontmatter (siehe vault_builder.py)",
    )
    parser.add_argument(
        "--embed",
        nargs="?",
        const="",
        default=None,
        metavar="MODELL",
        help="Nach Build+Index zusaetzlich Embeddings via Ollama berechnen "
        "(Modell optional, sonst ENV DOC2VAULT_EMBED_MODEL). Additiv: ist "
        "Ollama nicht erreichbar, laufen Konvertierung, Build und "
        "FTS5-Index trotzdem vollstaendig durch.",
    )
    parser.add_argument(
        "--duplicates",
        choices=["convert", "skip"],
        default="convert",
        help="Umgang mit inhaltsgleichen Quelldateien: convert = alle "
        "konvertieren und nur melden (Default), skip = je Gruppe nur die "
        "erste Datei konvertieren",
    )
    parser.add_argument(
        "--yes",
        "-y",
        action="store_true",
        help="Integrationsplan ohne Rueckfrage bestaetigen",
    )
    parser.add_argument(
        "--error-log",
        default=None,
        help="Optionaler Pfad fuer ein JSON-Fehlerprotokoll",
    )
    args = parser.parse_args(argv)

    input_root = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    if not input_root.is_dir():
        parser.error(f"Quellordner existiert nicht: {input_root}")
    path_error = check_paths(input_root, output_dir)
    if path_error:
        parser.error(path_error)
    if args.on_success == "archive" and not args.archive_dir:
        parser.error("--on-success archive erfordert --archive-dir")

    # Zielordner analysieren und Integrationsplan ableiten.
    profile = analyze_vault(output_dir)
    config = recommend_config(profile)
    config.do_ocr = args.ocr
    config.ocr_engine = args.ocr_engine
    config.ocr_languages = args.ocr_langs
    config.generate_picture_images = not args.no_images
    config.images_scale = args.images_scale
    config.do_table_structure = not args.no_tables
    config.xlsx_sheet_limit = args.xlsx_sheet_limit
    config.xlsx_on_limit = args.xlsx_on_limit
    config.on_success = args.on_success
    config.archive_dir = (
        str(Path(args.archive_dir).resolve()) if args.archive_dir else None
    )
    if args.notes_subdir is not None:
        config.notes_subdir = args.notes_subdir
    if args.attachments_subdir is not None:
        config.attachments_subdir = args.attachments_subdir
        config.attachments_mode = "central"
    if args.no_frontmatter:
        config.add_frontmatter = False

    engine_warning = check_ocr_engine(config)
    if engine_warning:
        print(f"FEHLER: {engine_warning}", file=sys.stderr)
        return 2

    # Ziel- und Archivordner ausschliessen: liegt einer davon in der Quelle,
    # wuerden erzeugte Notizen bzw. archivierte Originale beim naechsten Lauf
    # selbst als Quelldokumente eingesammelt.
    files = discover_files(
        input_root, exclude_dirs=(output_dir, config.archive_dir)
    )
    total = len(files)
    if total == 0:
        print(f"Keine unterstuetzten Dateien in {input_root} gefunden.")
        return 0

    # Duplikate (inhaltsgleiche Dateien) erkennen -- nicht zu verwechseln mit
    # der Idempotenz (unveraenderte Datei beim Wiederholungslauf).
    duplicate_groups = find_duplicate_files(files)
    skipped_duplicates: list[Path] = []
    if duplicate_groups and args.duplicates == "skip":
        for paths in duplicate_groups.values():
            # Deterministisch die erste (sortierte) Datei behalten.
            keep, *rest = sorted(paths)
            skipped_duplicates.extend(rest)
        skip_set = {str(p) for p in skipped_duplicates}
        files = [f for f in files if str(f) not in skip_set]
        total = len(files)

    # Plan EINMAL fuer den gesamten Batch anzeigen und bestaetigen lassen.
    print("\n=== Zielordner-Analyse ===")
    for obs in profile.observations:
        print(f"  • {obs}")
    print("\n=== Integrationsplan ===")
    for line in describe_plan(profile, config):
        print(f"  {line}")
    print(f"  Zu konvertieren: {total} Datei(en)")
    if duplicate_groups:
        dup_files = sum(len(p) for p in duplicate_groups.values())
        if args.duplicates == "skip":
            print(f"  Duplikate: {len(duplicate_groups)} Gruppe(n), "
                  f"{len(skipped_duplicates)} inhaltsgleiche Datei(en) "
                  "werden uebersprungen (--duplicates skip).")
        else:
            print(f"  Duplikate: {len(duplicate_groups)} Gruppe(n) mit "
                  f"{dup_files} inhaltsgleichen Dateien gefunden -- alle "
                  "werden konvertiert (--duplicates skip zum Ueberspringen).")
    if not config.do_ocr and any(
        Path(f).suffix.lower() in IMAGE_INPUT_EXTENSIONS for f in files
    ):
        print("  WARNUNG: Bilddateien im Batch, aber OCR ist aus -- "
              "gescannte Bilder ergeben leere Notizen (--ocr aktivieren).")
    if config.do_ocr and args.workers > 2:
        print(f"  HINWEIS: OCR mit {args.workers} parallelen Prozessen "
              "braucht viel RAM (je Prozess ein eigener Modellstapel). "
              "Bei Speicherfehlern (std::bad_alloc) -w 1 oder -w 2 nutzen.")
    if args.build_vault:
        print("  Vault-Build: Notizen → Inbox/, Bilder → Attachments/, "
              "Wikilinks + Frontmatter")
    if not args.yes:
        try:
            answer = input("\nSo einordnen und starten? [j/N] ").strip().lower()
        except EOFError:
            answer = ""
        if answer not in ("j", "ja", "y", "yes"):
            print("Abgebrochen.")
            return 0

    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n{total} Dateien gefunden. Starte mit {args.workers} Prozess(en)...")
    ok = 0
    reduced_count = 0
    failed: list[ConversionResult] = []
    start = time.perf_counter()

    def _cli_progress(done: int, total_n: int, res: ConversionResult) -> None:
        nonlocal ok, reduced_count
        if res.success:
            ok += 1
            if res.reduced_mode:
                reduced_count += 1
        else:
            failed.append(res)
        elapsed = time.perf_counter() - start
        rate = done / elapsed if elapsed else 0
        eta = (total_n - done) / rate if rate else 0
        marker = "  [reduziert]" if res.reduced_mode else ""
        if res.pdf_backend:
            marker += "  [pypdfium]"
        print(
            f"[{done}/{total_n}] ok={ok} fehler={len(failed)} "
            f"ETA={eta:6.0f}s  {Path(res.source_path).name}{marker}",
            flush=True,
        )

    try:
        run_conversion_batch(
            files, config, output_dir, input_root, args.workers,
            progress=_cli_progress,
        )
    except KeyboardInterrupt:
        print("\nAbgebrochen. Bereits konvertierte Dateien bleiben erhalten.")
        return 130

    print(f"\nFertig: {ok} erfolgreich, {len(failed)} fehlgeschlagen.")
    if reduced_count:
        print(f"  Davon {reduced_count} mit reduzierten Einstellungen "
              "(riesige Seiten, z. B. CAD-Plaene: Bildskalierung 1.0, "
              "ohne Bildextraktion).")
    if failed:
        by_cat: dict[str, int] = {}
        for r in failed:
            by_cat[r.error_category or "fehler"] = by_cat.get(r.error_category or "fehler", 0) + 1
        print("Fehler nach Kategorie:")
        for cat, count in sorted(by_cat.items(), key=lambda kv: -kv[1]):
            print(f"  {cat}: {count}")
    if failed and args.error_log:
        Path(args.error_log).write_text(
            json.dumps([asdict(r) for r in failed], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Fehlerprotokoll: {args.error_log}")

    if args.build_vault:
        # Lazy-Import: der reine Convert-Modus bleibt ohne python-frontmatter
        # lauffaehig.
        import vault_builder
        import vault_index

        print("\n=== Vault-Build ===")
        # Nur den frisch konvertierten Bereich bauen: liegt ein Import-
        # Unterordner vor (Standard bei bestehenden Vaults), bleiben die
        # uebrigen Notizen des Vaults unangetastet.
        build_source = (
            output_dir / config.notes_subdir if config.notes_subdir else output_dir
        )
        summary = vault_builder.build_vault(build_source, output_dir)
        print(f"  {summary.notes} Notiz(en) → Inbox/, "
              f"{summary.images} Bild(er) → Attachments/.")
        if summary.note_collisions or summary.image_collisions:
            print(f"  Kollisionen aufgeloest: {summary.note_collisions} "
                  f"Notiz(en), {summary.image_collisions} Bild(er).")

        idx = vault_index.update_index(output_dir)
        vault_index.write_index_md(output_dir)
        print(f"  Such-Index: {idx.indexed} neu/geändert, "
              f"{idx.total} Notizen insgesamt (INDEX.md aktualisiert).")

        if args.embed is not None:
            # Embeddings sind additiv: Fehler brechen den Lauf nicht ab.
            client = vault_index.OllamaClient()
            try:
                model = vault_index._resolve_model(
                    client, args.embed or None, "DOC2VAULT_EMBED_MODEL",
                    "Embeddings",
                )
                emb = vault_index.embed_vault(output_dir, client, model)
                print(f"  Embeddings: {emb.chunks_embedded} neu, "
                      f"{emb.chunks_reused} wiederverwendet "
                      f"(Modell {emb.model}, Dimension {emb.dimension}).")
            except vault_index.OllamaError as exc:
                print(f"  WARNUNG Embeddings uebersprungen: {exc}",
                      file=sys.stderr)

    return 1 if failed else 0


def main() -> int:
    """Einstiegspunkt fuer den ``doc2vault``-Konsolenbefehl."""
    return _run_cli()


if __name__ == "__main__":
    raise SystemExit(main())
